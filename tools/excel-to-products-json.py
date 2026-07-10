from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PRODUCT_SHEETS = ["Portionssnus", "Lössnus", "Aromer", "Tillbehör"]
EMPTY_VALUES = {None, ""}


def clean_value(value: Any) -> Any:
    if value in EMPTY_VALUES:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def sheet_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    products: list[dict[str, Any]] = []

    for sheet_name in PRODUCT_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]
        headers = [clean_value(cell.value) for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {
                str(header): clean_value(value)
                for header, value in zip(headers, row)
                if header and clean_value(value) is not None
            }
            if not record.get("product_id"):
                continue
            if str(record.get("visible_on_site", "Yes")).lower() == "no":
                continue
            record["_sheet"] = sheet_name
            products.append(record)

    return products


def main() -> None:
    parser = argparse.ArgumentParser(description="Export Swedsnus Excel product data to JSON for the static site.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out", type=Path, default=Path("data/products.json"))
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()

    products = sheet_rows(args.workbook)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(
        json.dumps(products, ensure_ascii=False, indent=2 if args.pretty else None, separators=None if args.pretty else (",", ":")),
        encoding="utf-8",
    )
    print(f"Exported {len(products)} rows to {args.out}")


if __name__ == "__main__":
    main()
